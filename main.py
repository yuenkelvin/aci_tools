#!/usr/bin/env python3

from openpyxl import load_workbook
import json


def read_ws(worksheet_name):
  ### read excel worksheet and return variable. 
  ws = wb[worksheet_name]
  rows = ws.max_row
  cols = ws.max_column

  header = dict((i, ws.cell(row=1, column=i).value) for i in range(1, cols+1))
  ws_dict = [ dict((header[i], ws.cell(row=j, column=i).value) for i in range (1,cols+1))  for j in range (2, rows+1) ] 

  print("max row: {}, max column: {}".format(rows,cols)) 

  return ws_dict


if __name__ == '__main__': 

  wb = load_workbook('config.xlsx')

  test_dict = read_ws('page2')
  print(json.dumps(test_dict,indent=4))

