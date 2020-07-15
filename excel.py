#!/usr/bin/env python3


import sys
import json
from openpyxl import load_workbook

def read_ws(workbook, worksheet):
  ### read excel worksheet and return variable. 
  wb = load_workbook(workbook)
  ws = wb[worksheet]
  rows = ws.max_row
  cols = ws.max_column

  header = dict((i, ws.cell(row=1, column=i).value) for i in range(1, cols+1))
  ws_dict = [ dict((header[i], ws.cell(row=j, column=i).value) for i in range (1,cols+1))  for j in range (2, rows+1) ] 

  print("max row: {}, max column: {}".format(rows,cols)) 

  return ws_dict

def main():

  test_dict = read_ws('config.xlsx', 'page2')
  print(json.dumps(test_dict,indent=4))

if __name__ == '__main__': 
  sys.exit(main())

